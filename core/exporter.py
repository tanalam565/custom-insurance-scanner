from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime
import json
from pathlib import Path
from config import Config

class Exporter:
    """Handles data export to various formats"""
    
    def __init__(self):
        self.export_folder = Path(Config.EXPORT_FOLDER)
        self.export_folder.mkdir(exist_ok=True)
    
    def export_to_excel(self, data, filename=None):
        """Export extracted data to Excel
        
        Args:
            data: Dictionary or list of dictionaries with extracted data
            filename: Optional custom filename
            
        Returns:
            Path to exported file
        """
        if filename is None:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f'renters_data_{timestamp}.xlsx'
        
        filepath = self.export_folder / filename
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Renters Insurance Data"
        
        # Convert single dict to list
        if isinstance(data, dict):
            data = [data]
        
        # Define headers
        headers = [
            'Insurance Company',
            'Policy Number',
            'Date Prepared',
            'Insurer Name',
            'Insurer Address',
            'Insurer City State',
            'Insurance Amount',
            'Property Address',
            'Effective Date',
            'Expiration Date',
            'Extraction Date'
        ]
        
        # Write headers with styling
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Write data
        for row_idx, record in enumerate(data, start=2):
            ws.cell(row=row_idx, column=1, value=record.get('insurance_company', ''))
            ws.cell(row=row_idx, column=2, value=record.get('policy_number', ''))
            ws.cell(row=row_idx, column=3, value=record.get('date_prepared', ''))
            ws.cell(row=row_idx, column=4, value=record.get('insurer_name', ''))
            ws.cell(row=row_idx, column=5, value=record.get('insurer_address', ''))
            ws.cell(row=row_idx, column=6, value=record.get('insurer_city_state', ''))
            ws.cell(row=row_idx, column=7, value=record.get('insurance_amount', ''))
            ws.cell(row=row_idx, column=8, value=record.get('property_address', ''))
            ws.cell(row=row_idx, column=9, value=record.get('effective_date', ''))
            ws.cell(row=row_idx, column=10, value=record.get('expiration_date', ''))
            ws.cell(row=row_idx, column=11, value=datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
        
        # Adjust column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[chr(64 + col)].width = 20
        
        # Save workbook
        wb.save(filepath)
        
        return filepath
    
    def export_to_json(self, data, filename=None):
        """Export extracted data to JSON
        
        Args:
            data: Dictionary or list of dictionaries with extracted data
            filename: Optional custom filename
            
        Returns:
            Path to exported file
        """
        if filename is None:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f'renters_data_{timestamp}.json'
        
        filepath = self.export_folder / filename
        
        # Add metadata
        export_data = {
            'export_date': datetime.now().isoformat(),
            'record_count': len(data) if isinstance(data, list) else 1,
            'data': data
        }
        
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(export_data, f, indent=2, ensure_ascii=False)
        
        return filepath
    
    def export_to_csv(self, data, filename=None):
        """Export extracted data to CSV
        
        Args:
            data: Dictionary or list of dictionaries with extracted data
            filename: Optional custom filename
            
        Returns:
            Path to exported file
        """
        import csv
        
        if filename is None:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f'renters_data_{timestamp}.csv'
        
        filepath = self.export_folder / filename
        
        # Convert single dict to list
        if isinstance(data, dict):
            data = [data]
        
        # Define fieldnames
        fieldnames = [
            'insurance_company',
            'policy_number',
            'date_prepared',
            'insurer_name',
            'insurer_address',
            'insurer_city_state',
            'insurance_amount',
            'property_address',
            'effective_date',
            'expiration_date',
            'extraction_date'
        ]
        
        with open(filepath, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            
            for record in data:
                row = {key: record.get(key, '') for key in fieldnames[:-1]}
                row['extraction_date'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                writer.writerow(row)
        
        return filepath
    
    def create_batch_export(self, data_list, format='excel'):
        """Export multiple records in batch
        
        Args:
            data_list: List of dictionaries with extracted data
            format: Export format ('excel', 'json', 'csv')
            
        Returns:
            Path to exported file
        """
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'batch_export_{timestamp}'
        
        if format == 'excel':
            return self.export_to_excel(data_list, f'{filename}.xlsx')
        elif format == 'json':
            return self.export_to_json(data_list, f'{filename}.json')
        elif format == 'csv':
            return self.export_to_csv(data_list, f'{filename}.csv')
        else:
            raise ValueError(f"Unsupported format: {format}")